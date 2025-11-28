import pandas as pd
import numpy as np
import sys
import logging
import requests
import os
from sqlalchemy import create_engine
from sklearn.preprocessing import LabelEncoder, MinMaxScaler, StandardScaler
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
import statsmodels.api as sm
from matplotlib.backends.backend_pdf import PdfPages
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from sklearn.linear_model import LinearRegression, LogisticRegression
import smtplib
from email.message import EmailMessage
from sklearn.metrics import mean_squared_error, r2_score, accuracy_score, precision_score, recall_score, f1_score, roc_auc_score
import schedule
import time
import statsmodels.api as sm

def analyze_time_series(df, date_col, numeric_col, period=None):
    series = df.set_index(date_col)[numeric_col]
    # Удаляем NaN перед декомпозицией
    series = series.dropna()
    decomposition = sm.tsa.seasonal_decompose(series, model='additive', period=period)
    # остальной код...
# Настройка логирования
logging.basicConfig(level=logging.INFO, stream=sys.stdout,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Параметры базы данных для хранения результатов
DB_PARAMS = {
    'host': 'localhost',
    'port': 5432,
    'user': 'postgres',
    'password': 'kapibara',
    'dbname': 'demo'
}

# Создаем SQLAlchemy engine
def get_engine():
    conn_str = f"postgresql://{DB_PARAMS['user']}:{DB_PARAMS['password']}@{DB_PARAMS['host']}:{DB_PARAMS['port']}/{DB_PARAMS['dbname']}"
    return create_engine(conn_str)

# Функции для сохранения данных и моделей в базу
def save_analysis_result(table_name, df):
    try:
        engine = get_engine()
        df.to_sql(table_name, con=engine, if_exists='replace', index=False)
        logging.info(f"Результаты сохранены в таблицу '{table_name}'.")
    except Exception as e:
        logging.error(f"Ошибка при сохранении в таблицу '{table_name}': {e}")

def save_model_parameters(model, table_name, feature_cols):
    # Исправляем создание DataFrame с одинаковой длиной массивов
    try:
        engine = get_engine()
        coef = getattr(model, 'coef_', np.array([np.nan]))
        intercept = getattr(model, 'intercept_', np.nan)

        # Обеспечиваем массив одинаковой длины
        coef = np.array(coef)
        intercepts = np.full(shape=len(feature_cols), fill_value=intercept)

        if coef.ndim == 1:
            coefficients = coef
        else:
            coefficients = coef.flatten()

        if len(coefficients) != len(feature_cols):
            # В случае несоответствия, заполняем NaN
            coefficients = np.full(shape=len(feature_cols), fill_value=np.nan)

        params_df = pd.DataFrame({
            'feature': feature_cols,
            'coefficient': coefficients,
            'intercept': intercepts
        })
        params_df.to_sql(table_name, con=engine, if_exists='replace', index=False)
        logging.info(f"Модельные параметры сохранены в '{table_name}'.")
    except Exception as e:
        logging.error(f"Ошибка при сохранении модели в '{table_name}': {e}")

#  Визуализация функций
def plot_histogram(df, column):
    plt.figure(figsize=(8, 5))
    sns.histplot(df[column], kde=True)
    plt.title(f'Гистограмма: {column}')
    plt.xlabel(column)
    plt.ylabel('Частота')
    plt.tight_layout()

def plot_boxplot(df, column):
    plt.figure(figsize=(8, 5))
    sns.boxplot(x=df[column])
    plt.title(f'Boxplot: {column}')
    plt.tight_layout()

def plot_scatter(df, x_col, y_col):
    plt.figure(figsize=(8, 6))
    sns.scatterplot(data=df, x=x_col, y=y_col)
    plt.title(f'Диаграмма рассеяния: {x_col} vs {y_col}')
    plt.tight_layout()

def plot_correlation_heatmap(df, numeric_cols):
    plt.figure(figsize=(10,8))
    corr = df[numeric_cols].corr()
    sns.heatmap(corr, annot=True, fmt=".2f", cmap='coolwarm')
    plt.title('Тепловая карта корреляций')
    plt.tight_layout()

def plot_pairplot(df, cols):
    sns.pairplot(df[cols])
    plt.tight_layout()

def plot_interactive_plot(df, x, y, title=''):
    fig = px.scatter(df, x=x, y=y, title=title)
    fig.show()

# Вспомогательные функции
def load_csv(file_path, sep_list=[',', '\t', ';']):
    for sep in sep_list:
        try:
            df = pd.read_csv(file_path, sep=sep)
            logging.info(f"Успешная загрузка с разделителем '{sep}':")
            logging.info(df.head())
            return df
        except Exception as e:
            logging.warning(f"Не удалось загрузить с разделителем '{sep}': {e}")
    logging.error("Все попытки загрузки не удались.")
    return None

def load_excel(file_path, sheet_name=0):
    try:
        df_excel = pd.read_excel(file_path, sheet_name=sheet_name)
        logging.info("Excel данные загружены:")
        logging.info(df_excel.head())
        return df_excel
    except Exception as e:
        logging.error(f"Ошибка при загрузке Excel: {e}")
        return None

def validate_data(df, key_columns=None, numeric_columns=None, handle_missing='drop', fill_method='mean'):
    validation_report = {}

    # Проверка на дубликаты
    if key_columns:
        missing_keys = [col for col in key_columns if col not in df.columns]
        if missing_keys:
            logging.warning(f"Ключевые колонки отсутствуют: {missing_keys}")
        else:
            duplicates = df[df.duplicated(subset=key_columns)]
            validation_report['duplicates'] = len(duplicates)
            if len(duplicates) > 0:
                logging.warning(f"Обнаружены дубликаты по ключам {key_columns}: {len(duplicates)}")
                df = df.drop_duplicates(subset=key_columns)
            else:
                logging.info("Дубликаты по ключам отсутствуют.")

    # Проверка на пропуски
    missing = df.isnull().sum()
    validation_report['missing'] = missing
    logging.info("Количество пропусков по колонкам:")
    logging.info(missing)

    # Обработка пропусков
    if handle_missing == 'drop':
        before_rows = len(df)
        df = df.dropna()
        after_rows = len(df)
        validation_report['rows_dropped'] = before_rows - after_rows
        logging.info(f"Удалено строк с пропусками: {before_rows - after_rows}")
    elif handle_missing in ['mean', 'median']:
        for col in df.columns:
            if df[col].isnull().any():
                if pd.api.types.is_numeric_dtype(df[col]):
                    fill_value = df[col].mean() if handle_missing == 'mean' else df[col].median()
                    df[col] = df[col].fillna(fill_value)
                    logging.info(f"Заполнены пропуски в '{col}' значением {fill_method}: {fill_value}")
    else:
        logging.warning(f"Неизвестный метод обработки пропусков: {handle_missing}")

    # Обработка выбросов - IQR
    if numeric_columns:
        for col in numeric_columns:
            if col in df.columns:
                # Приведение к числовому типу
                df[col] = pd.to_numeric(df[col], errors='coerce')
                Q1 = df[col].quantile(0.25)
                Q3 = df[col].quantile(0.75)
                IQR = Q3 - Q1
                lower_bound = Q1 - 1.5 * IQR
                upper_bound = Q3 + 1.5 * IQR
                # Заменяем только выбросы через clip
                df[col] = df[col].clip(lower=lower_bound, upper=upper_bound)
                logging.info(f"Обработаны выбросы в '{col}' по IQR.")
    # Обработка выбросов - Z-score
    if numeric_columns:
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
                mean = df[col].mean()
                std = df[col].std()
                if std != 0:
                    z_scores = (df[col] - mean) / std
                    # Заменяем только выбросы на среднее
                    df[col] = np.where(z_scores.abs() > 3, mean, df[col])
                    logging.info(f"Обработаны выбросы в '{col}' по Z-score.")
    return df, validation_report

def scale_numeric_features(df, numeric_columns, method='minmax'):
    scaler = None
    if method == 'minmax':
        scaler = MinMaxScaler()
    elif method == 'standard':
        scaler = StandardScaler()
    else:
        logging.warning(f"Неизвестный метод масштабирования: {method}")
        return df

    for col in numeric_columns:
        if col in df.columns:
            # Приведение к числовому типу
            df[col] = pd.to_numeric(df[col], errors='coerce')
    scaled_values = scaler.fit_transform(df[numeric_columns])
    df_scaled = df.copy()
    df_scaled[numeric_columns] = scaled_values
    logging.info(f"Масштабирование числовых признаков выполнено методом '{method}'.")
    return df_scaled

def query_postgresql(db_params, sql_query):
    try:
        conn_str = f"postgresql://{db_params['user']}:{db_params['password']}@{db_params['host']}:{db_params['port']}/{db_params['dbname']}"
        engine = create_engine(conn_str)
        with engine.connect() as connection:
            df_db = pd.read_sql(sql_query, connection)
        logging.info("Данные из базы данных PostgreSQL:")
        logging.info(df_db.head())
        return df_db
    except Exception as e:
        logging.error(f"Ошибка при подключении или выполнении запроса к PostgreSQL: {e}")
        return None

def load_from_api(api_url, params=None):
    try:
        response = requests.get(api_url, params=params)
        if response.status_code == 200:
            response.encoding = 'utf-8'
            data = response.json()
            df_api = pd.DataFrame(data)
            logging.info("Данные из API получены:")
            logging.info(df_api.head())
            return df_api
        else:
            logging.error(f"Ошибка при запросе API: {response.status_code}")
            return None
    except Exception as e:
        logging.error(f"Ошибка при обращении к API: {e}")
        return None

def encode_categorical(df, categorical_columns, method='onehot'):
    df_encoded = df.copy()
    if method == 'onehot':
        df_encoded = pd.get_dummies(df_encoded, columns=categorical_columns)
        logging.info(f"Применено One-Hot Encoding к: {categorical_columns}")
    elif method == 'label':
        le = LabelEncoder()
        for col in categorical_columns:
            if col in df_encoded.columns:
                df_encoded[col] = le.fit_transform(df_encoded[col].astype(str))
                logging.info(f"Применено Label Encoding к: {col}")
    else:
        logging.warning(f"Неизвестный метод кодирования: {method}")
    return df_encoded

def get_categorical_columns(df, threshold=0.05):
    categorical_cols = []
    for col in df.columns:
        if pd.api.types.is_object_dtype(df[col]) or isinstance(df[col].dtype, pd.CategoricalDtype):
            unique_ratio = df[col].nunique() / len(df)
            if unique_ratio < threshold:
                categorical_cols.append(col)
    return categorical_cols

def calculate_statistics(df, numeric_columns):
    stats = {}
    for col in numeric_columns:
        if col in df.columns:
            stats[col] = {
                'mean': df[col].mean(),
                'median': df[col].median(),
                'mode': df[col].mode()[0] if not df[col].mode().empty else None,
                'std_dev': df[col].std()
            }
    return stats

def analyze_time_series(df, date_col, target_col, period=None):
    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
    df = df.dropna(subset=[date_col])
    df.set_index(date_col, inplace=True)
    df = df.sort_index()

    series = df[target_col]

    if len(series) >= 2:
        if period is None:
            period = pd.infer_freq(series.index)
            if period is None:
                period = 12
        decomposition = sm.tsa.seasonal_decompose(series, model='additive', period=period)
        decomposition.plot()
        plt.suptitle(f'Декомпозиция: {target_col}')
        plt.tight_layout()

        # Обнаружение аномалий
        mean = series.mean()
        std = series.std()
        z_scores = (series - mean) / std
        anomalies = series[z_scores.abs() > 3]
        plt.figure(figsize=(14,6))
        plt.plot(series.index, series, label='Исходные данные')
        plt.scatter(anomalies.index, anomalies, color='red', label='Аномалии')
        plt.title(f'Анализ временного ряда: {target_col}')
        plt.xlabel('Дата')
        plt.ylabel(target_col)
        plt.legend()
        plt.tight_layout()
        plt.show()

        logging.info(f"Обнаружено {len(anomalies)} аномалий в {target_col}.")
    else:
        logging.warning("Недостаточно данных для анализа временного ряда.")
    return

def train_regression_model(df, feature_cols, target_col, test_size=0.2, random_state=42):
    data = df[feature_cols + [target_col]].dropna()
    X = data[feature_cols]
    y = data[target_col]
    X = X.copy()
    y = y.copy()
    # Обязательно приводим к float
    X = X.apply(pd.to_numeric, errors='coerce')
    y = pd.to_numeric(y, errors='coerce')
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=test_size, random_state=random_state)
    model = LinearRegression()
    model.fit(X_train, y_train)
    y_pred = model.predict(X_test)
    mse = mean_squared_error(y_test, y_pred)
    rmse = np.sqrt(mse)
    r2 = r2_score(y_test, y_pred)
    print(f"Регрессия - MSE: {mse:.4f}, RMSE: {rmse:.4f}, R2: {r2:.4f}")
    # Сохраняем результаты в БД
    result_df = pd.DataFrame({
        'model_type': ['LinearRegression'],
        'mse': [mse],
        'rmse': [rmse],
        'r2': [r2],
        'features': [', '.join(feature_cols)]
    })
    save_analysis_result('regression_results', result_df)
    # Сохраняем параметры модели
    save_model_parameters(model, 'regression_model_params', feature_cols)
    return model

def train_classification_model(df, feature_cols, target_col, test_size=0.2, random_state=42):
    X = df[feature_cols]
    y = df[target_col]
    X = X.copy()
    y = y.copy()
    # Приведение к числовому типу
    X = X.apply(pd.to_numeric, errors='coerce')
    y = pd.to_numeric(y, errors='coerce')
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=test_size, random_state=random_state)

    model = LogisticRegression(max_iter=1000)
    model.fit(X_train, y_train)
    y_pred = model.predict(X_test)
    y_proba = model.predict_proba(X_test)

    accuracy = accuracy_score(y_test, y_pred)
    precision = precision_score(y_test, y_pred, average='weighted', zero_division=0)
    recall = recall_score(y_test, y_pred, average='weighted', zero_division=0)
    f1 = f1_score(y_test, y_pred, average='weighted', zero_division=0)

    try:
        roc_auc = roc_auc_score(y_test, y_proba, multi_class='ovr', average='weighted')
    except:
        roc_auc = None

    # Сохраняем результаты
    result_df = pd.DataFrame({
        'model_type': ['LogisticRegression'],
        'accuracy': [accuracy],
        'precision': [precision],
        'recall': [recall],
        'f1_score': [f1],
        'roc_auc': [roc_auc],
        'features': [', '.join(feature_cols)]
    })
    save_analysis_result('classification_results', result_df)
    save_model_parameters(model, 'classification_model_params', feature_cols)

    print(f"Классификация - Accuracy: {accuracy:.4f}")
    print(f"Precision: {precision:.4f}")
    print(f"Recall: {recall:.4f}")
    print(f"F1-score: {f1:.4f}")
    if roc_auc is not None:
        print(f"ROC-AUC: {roc_auc:.4f}")
    else:
        print("ROC-AUC не может быть рассчитан.")
    return model

#  Функция отправки email
def send_email_with_attachment(smtp_server, smtp_port, login, password, subject, body, to_email, attachment_path):
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = login
    msg['To'] = to_email
    msg.set_content(body)

    # Чтение файла
    with open(attachment_path, 'rb') as f:
        file_data = f.read()
        file_name = os.path.basename(attachment_path)
    msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(login, password)
            server.send_message(msg)
        logging.info("Отчет успешно отправлен по email.")
    except Exception as e:
        logging.error(f"Ошибка при отправке email: {e}")

# Основной блок
def main_task():
    logging.info("Запуск основной задачи обработки данных.")
    # Пути и параметры
    file_path = 'E:/test.csv'
    excel_path = 'E:/test.xlsx'
    api_url = 'https://jsonplaceholder.typicode.com/posts'

    # Email параметры
    SMTP_SERVER = 'smtp.yandex.ru'
    SMTP_PORT = 587
    EMAIL_LOGIN = 'ale1242@yandex.ru'
    EMAIL_PASSWORD = '_____________'  # вставьте ваш реальный пароль
    EMAIL_TO = 'ale1242@yandex.ru'

    # Глобальные переменные для отчета
    global csv_data, csv_valid, csv_scaled, csv_encoded
    global excel_data, excel_valid, excel_scaled, excel_encoded

    # Загрузка CSV
    if os.path.exists(file_path):
        csv_data = load_csv(file_path, sep_list=[';'])
        if csv_data is not None:
            # Обработка дат
            if 'first_air_date' in csv_data.columns:
                csv_data['first_air_date'] = pd.to_datetime(csv_data['first_air_date'], errors='coerce')
            if 'last_air_date' in csv_data.columns:
                csv_data['last_air_date'] = pd.to_datetime(csv_data['last_air_date'], errors='coerce')

            csv_data.columns = csv_data.columns.str.strip().str.lower()

            # Визуализации
            if 'vote_count' in csv_data.columns:
                plot_histogram(csv_data, 'vote_count')
                plot_boxplot(csv_data, 'vote_count')

            if 'first_air_date' in csv_data.columns:
                analyze_time_series(csv_data.copy(), 'first_air_date', 'vote_count')

            # Категориальные колонки
            categorical_cols = get_categorical_columns(csv_data)
            logging.info("Категориальные колонки: %s", categorical_cols)

            # Валидация
            csv_valid, report_csv = validate_data(
                csv_data,
                key_columns=['id'],
                numeric_columns=['vote_count', 'vote_average', 'popularity'],
                handle_missing='mean'
            )

            # Визуализация
            plot_correlation_heatmap(csv_valid, ['vote_count', 'vote_average', 'popularity'])

            # Масштабирование
            csv_scaled = scale_numeric_features(csv_valid, ['vote_count', 'vote_average', 'popularity'], method='standard')

            # Кодирование
            csv_encoded = encode_categorical(csv_scaled, categorical_cols, method='onehot')
            logging.info("Обработанные данные (пример):")
            logging.info(csv_encoded.head())

            # Визуализация scatter и интерактив
            plot_scatter(csv_encoded, 'vote_average', 'popularity')
            plot_interactive_plot(csv_encoded, 'vote_average', 'popularity', 'Vote Average vs Popularity')

            # Обучение моделей
            if 'vote_count' in csv_encoded.columns:
                logging.info("Обучение регрессии:")
                model_reg = train_regression_model(csv_encoded, ['vote_average', 'popularity'], 'vote_count')

            if 'type' in csv_encoded.columns:
                csv_for_classification = csv_encoded.dropna(subset=['type'])
                if len(csv_for_classification) > 1:
                    logging.info("Обучение классификации:")
                    model_clf = train_classification_model(csv_for_classification, ['vote_average', 'popularity'], 'type')
                else:
                    logging.info("Недостаточно данных для классификации по 'type'")
    else:
        logging.warning(f"Файл {file_path} не найден.")

    # Аналогично для Excel
    if os.path.exists(excel_path):
        excel_data = load_excel(excel_path)
        if excel_data is not None:
            if 'first_air_date' in excel_data.columns:
                excel_data['first_air_date'] = pd.to_datetime(excel_data['first_air_date'], errors='coerce')
            if 'last_air_date' in excel_data.columns:
                excel_data['last_air_date'] = pd.to_datetime(excel_data['last_air_date'], errors='coerce')

            excel_data.columns = excel_data.columns.str.strip().str.lower()

            if 'vote_count' in excel_data.columns:
                plot_histogram(excel_data, 'vote_count')
                plot_boxplot(excel_data, 'vote_count')

            if 'first_air_date' in excel_data.columns:
                analyze_time_series(excel_data.copy(), 'first_air_date', 'vote_count')

            categorical_cols_excel = get_categorical_columns(excel_data)
            logging.info("Категориальные колонки Excel: %s", categorical_cols_excel)

            excel_valid, report_excel = validate_data(
                excel_data,
                key_columns=['id'],
                numeric_columns=['number_of_seasons', 'number_of_episodes', 'vote_count', 'vote_average', 'popularity'],
                handle_missing='mean'
            )

            plot_correlation_heatmap(excel_valid, ['vote_count', 'vote_average', 'popularity'])

            excel_scaled = scale_numeric_features(excel_valid, ['number_of_seasons', 'number_of_episodes', 'vote_count', 'vote_average', 'popularity'], method='standard')

            excel_encoded = encode_categorical(excel_scaled, categorical_cols_excel, method='onehot')
            logging.info("Обработанные данные Excel (пример):")
            logging.info(excel_encoded.head())

            plot_scatter(excel_encoded, 'vote_average', 'popularity')
            plot_interactive_plot(excel_encoded, 'vote_average', 'popularity', 'Vote Average vs Popularity')

            if 'vote_count' in excel_encoded.columns:
                logging.info("Обучение регрессии (Excel):")
                model_reg_excel = train_regression_model(excel_encoded, ['vote_average', 'popularity'], 'vote_count')

            if 'type' in excel_encoded.columns:
                excel_for_classification = excel_encoded.dropna(subset=['type'])
                if len(excel_for_classification) > 1:
                    train_classification_model(excel_for_classification, ['vote_average', 'popularity'], 'type')
                else:
                    logging.info("Недостаточно данных для классификации по 'type'")
    else:
        logging.warning(f"Файл {excel_path} не найден.")

    # Работа с базой данных
    db_params = {
        'host': 'localhost',
        'port': 5432,
        'user': 'postgres',
        'password': 'kapibara',
        'dbname': 'demo'
    }
    sql_query = "SELECT * FROM boarding_passes LIMIT 100"
    db_data = query_postgresql(db_params, sql_query)
    if db_data is not None:
        save_analysis_result('db_query_result', db_data)

    # Работа с API
    api_params = {}
    api_url = 'https://jsonplaceholder.typicode.com/posts'
    api_data = load_from_api(api_url, params=api_params)
    if api_data is not None:
        save_analysis_result('api_data', api_data)

    # Вызов функции сбора отчета с форматированием Excel и отправкой по email
    def create_full_report_and_send():
        report_path = r"C:\Users\a.kislicin\Reports\full_report.pdf"
        os.makedirs(os.path.dirname(report_path), exist_ok=True)
        with PdfPages(report_path) as pdf:
            # Визуализации по CSV
            if 'csv_data' in globals() and csv_data is not None:
                plt.figure()
                sns.histplot(csv_data, y='vote_count')
                plt.title('Гистограмма vote_count')
                pdf.savefig()
                plt.close()

                plt.figure()
                sns.boxplot(x=csv_data['vote_count'])
                plt.title('Boxplot vote_count')
                pdf.savefig()
                plt.close()

                plt.figure()
                sns.scatterplot(data=csv_data, x='vote_average', y='popularity')
                plt.title('Vote Average vs Popularity')
                pdf.savefig()
                plt.close()

                if 'first_air_date' in csv_data.columns and 'vote_count' in csv_data.columns:
                    plt.figure()
                    series = csv_data['vote_count']
                    series.plot()
                    plt.title('Time series vote_count')
                    plt.tight_layout()
                    pdf.savefig()
                    plt.close()

            # Визуализации по Excel
            if 'excel_data' in globals() and excel_data is not None:
                plt.figure()
                sns.histplot(excel_data, y='vote_count')
                plt.title('Гистограмма vote_count (Excel)')
                pdf.savefig()
                plt.close()

                plt.figure()
                sns.boxplot(x=excel_data['vote_count'])
                plt.title('Boxplot vote_count (Excel)')
                pdf.savefig()
                plt.close()

                plt.figure()
                sns.scatterplot(data=excel_data, x='vote_average', y='popularity')
                plt.title('Vote Average vs Popularity (Excel)')
                pdf.savefig()
                plt.close()

                if 'first_air_date' in excel_data.columns and 'vote_count' in excel_data.columns:
                    plt.figure()
                    series = excel_data['vote_count']
                    series.plot()
                    plt.title('Time series vote_count (Excel)')
                    plt.tight_layout()
                    pdf.savefig()
                    plt.close()

        # Форматирование таблиц Excel перед сохранением
        # Форматируем CSV DataFrame
        if 'csv_data' in globals() and csv_data is not None:
            csv_path_formatted = r"C:\Users\a.kislicin\Reports\csv_formatted.xlsx"
            with pd.ExcelWriter(csv_path_formatted, engine='openpyxl') as writer:
                csv_data.to_excel(writer, index=False, sheet_name='CSV Data')
                wb = writer.book
                ws = wb['CSV Data']
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max(max_length + 2, 10)
                    ws.column_dimensions[col_letter].width = adjusted_width

        # Форматируем Excel DataFrame
        if 'excel_data' in globals() and excel_data is not None:
            excel_path_formatted = r"C:\Users\a.kislicin\Reports\excel_formatted.xlsx"
            with pd.ExcelWriter(excel_path_formatted, engine='openpyxl') as writer:
                excel_data.to_excel(writer, index=False, sheet_name='Excel Data')
                wb2 = writer.book
                ws2 = wb2['Excel Data']
                for cell in ws2[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                for col in ws2.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max(max_length + 2, 10)
                    ws2.column_dimensions[col_letter].width = adjusted_width

        # Отправка по email
        send_email_with_attachment(
            smtp_server=SMTP_SERVER,
            smtp_port=SMTP_PORT,
            login=EMAIL_LOGIN,
            password=EMAIL_PASSWORD,
            subject='Отчет по данным',
            body='В приложении полный отчет.',
            to_email=EMAIL_TO,
            attachment_path=report_path
        )

    # Вызов сборки и отправки
    create_full_report_and_send()
    logging.info("Завершение основной задачи обработки данных.")

#  Запуск планировщика
def schedule_tasks():
    logging.info("Запуск планировщика задач.")
    # Настройка расписания: например, запускать раз в день в 18:47
    schedule.every().day.at("18:47").do(main_task)
    while True:
        schedule.run_pending()
        time.sleep(60)  # проверять каждую минуту

if __name__ == "__main__":
    try:
        schedule_tasks()
    except KeyboardInterrupt:
        logging.info("Планировщик остановлен пользователем.")